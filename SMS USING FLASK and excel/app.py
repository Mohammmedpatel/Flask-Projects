from flask import Flask, jsonify, request, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
import openpyxl

app=Flask(__name__)

app.config["SQLALCHEMY_DATABASE_URI"]="sqlite:///students.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"]= False
db=SQLAlchemy(app)
ma=Marshmallow(app)

class Student(db.Model):
    id=db.Column(db.Integer, primary_key=True)
    name=db.Column(db.String(30))
    age=db.Column(db.Integer)
    gender=db.Column(db.String(10))
    email=db.Column(db.String(100),unique=True)
    phone=db.Column(db.String(10))
    course=db.Column(db.String(20))
    enrollment_date=db.Column(db.String(20))
    city=db.Column(db.String(20))
    grade=db.Column(db.String(3))

class StudentsSchema(ma.SQLAlchemyAutoSchema):
    class Meta:
        model = Student
        fields= ["id","name","age","gender","email","phone","course","enrollment_date","city","grade"]
        # load_instance=True

student_schema=StudentsSchema()
students_schema=StudentsSchema(many=True)

with app.app_context():
    db.create_all()

@app.route("/")
def home():
    return "hello"    

@app.route("/students", methods=["POST"])
def add_student():
    data=request.get_json()
    new_data=student_schema.load(data)
    student=Student(**new_data)
    db.session.add(student)
    db.session.commit()
    return jsonify({"message":"Successfully Added"})

@app.route("/students", methods=["GET"])
def get_student():
    students=Student.query.all()
    return students_schema.dump(students)

@app.route("/students/import", methods=["POST"])
def import_excel():
    file=request.files["file"]
    wb=openpyxl.load_workbook(file)
    sheet=wb.active
    for row in sheet.iter_rows(min_row=2,values_only=True):
        print(row)
        new_student=Student(name=row[0],age=row[1],gender=row[2],email=row[3],phone=row[4],course=row[5],enrollment_date=row[6],city=row[7],grade=row[8])
        db.session.add(new_student)
    db.session.commit()
    return "Excel imported"

@app.route("/students/export", methods=["GET"])
def export_excel():
    students=Student.query.all()
    wb=openpyxl.Workbook()
    sheet=wb.active
    sheet.append(["Name","Age","Gender","Email","Phone","Course","Enrollment_date","City","Grade"])
    for student in students:
        sheet.append([student.name,student.age,student.email,student.phone,student.course,student.enrollment_date,student.city,student.grade])
    filename="student.xlsx"

    wb.save(filename)
    return send_file(filename, as_attachment=True)    

if __name__=="__main__":
    app.run(debug=True)
